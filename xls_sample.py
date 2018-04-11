import os
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.comments import Comment
import random
from collections import defaultdict

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
MEATADATA_CELL = 'A1'


def gen_dict(a):
    d = defaultdict(list)
    # the score would be look like this [('image_name', response_time)...]
    # Group the list by image_name and combine the score
    for k, v in a:
        d[k].append(v)
    # Output: defaultdict(<type 'list'>, {'a': [1, 2, 3], 'c': [1, 2, 3], 'b': [1, 2, 3]})
    # >>> d['a']
    # [1, 2, 3]
    # We flatten it into a list of list and later save it as csv
    score_list = []
    for name, score in d.iteritems():
        x = [name]
        x.extend(score)
        score_list.append(x)
    # >>> score_list
    # [['a', 1, 2, 3], ['c', 1, 2, 3], ['b', 1, 2, 3]]
    score_list.sort(key=lambda x: x[0])
    return score_list


def get_random_score(min, max):
    return random.randint(min, max) + random.random()


def generate_score():
    images = 'ABCDEFGHIJKLMN'
    image_set = list(images) * 12
    random.shuffle(image_set)
    scores = [(get_random_score(20, 50), get_random_score(100, 500)) for x in image_set]
    result = zip(image_set, scores)
    print 'ORIGINAL RESULT:'
    print result
    return gen_dict(result)


def ws_append_data(ws, row_start, col_start, data):
    curr_col = col_start
    curr_row = row_start
    for d in data:
        if isinstance(d, tuple):
            ws.cell(column=curr_col, row=curr_row, value=d[0]).border = thin_border
            ws.cell(column=curr_col+1, row=curr_row, value=d[1]).border = thin_border
            curr_col += 2
        else:
            ws.cell(column=curr_col, row=curr_row, value=d).border = thin_border
            curr_col += 1


def generate_random_identity():
    no = str(int(time.time()))
    name = random.choice(['Ayi', 'Nia', 'Yopi', 'Utari', 'Fauzan', 'Rola'])
    gender = random.choice(['L', 'P'])
    origin = random.choice(['X', 'Y'])
    age = random.randint(12, 15)
    return [no, name, gender, age, origin]


def save(data, sesi):
    fn = 'hasil/output.xlsx'
    if os.path.exists(fn):
        wb = load_workbook(fn)
    else:
        wb = load_workbook('hasil/_output_template.xlsx')
    if os.path.exists('hasil/~$output.xlsx'):
        raise IOError('File output is locked!')
    ws1 = wb.active  # worksheet 1
    ws1.title = 'Response'
    comment = ws1[MEATADATA_CELL].comment
    row_start = 4 if not comment else int(comment.text.split(':')[1])
    if sesi == 1:
        ws_append_data(ws1, row_start, 1, generate_random_identity())
    else:
        ws_append_data(ws1, row_start, 1, [''] * 6)
    ws_append_data(ws1, row_start, 6, ['Nama Blok'])
    ws_append_data(ws1, row_start, 7, data[0])
    for response in data[1:]:
        row_start += 1
        ws_append_data(ws1, row_start, 1, [''] * 6)
        ws_append_data(ws1, row_start, 7, response)
    msg = '[COMMENT INI JANGAN DIAPA-APAIN!]:%d' % (row_start + 1)
    comment = Comment(msg, 'METADATA')
    ws1[MEATADATA_CELL].comment = comment
    wb.save(filename=fn)


def simulate_session():
    for x in range(3):
        data = generate_score()
        save(data, x+1)


if __name__ == "__main__":
    for x in range(50):
        simulate_session()
